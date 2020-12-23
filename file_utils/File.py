import json
import openpyxl
import urllib

class HTML:
    def save(text,path):
        with open(path, 'w+') as f:
            f.write(text)
            f.close()

    def read(path):
        page = urllib.urlopen(path).read()
        return page


class JSON: 
    def save(text, path):
        s = json.dumps(text, indent=4, ensure_ascii=False)
        with open(path, "w+") as f:
            f.write(s)
            f.close()
        f.close()

    def read(path):
        with open(path, 'r') as f:
            data = json.load(f)
            f.close()
            return data

class Excel_Jyxx:
    def save_toExcel(text,path, wk_sheet="main"):
        
        # open workbook
        workbook = openpyxl.load_workbook(path)

        if(wk_sheet in workbook.sheetnames): 
            ws = workbook[wk_sheet]
            workbook.remove(ws)
        
        if('Placeholder' in workbook.sheetnames): 
            ws = workbook['Placeholder']
            workbook.remove(ws)
        
        worksheet = workbook.create_sheet(wk_sheet)
        worksheet.column_dimensions['A'].width = 30
        worksheet.column_dimensions['B'].width = 120
        worksheet.column_dimensions['C'].width = 120
        # worksheet.column_dimensions[ 0 ].height= 30

        # Attempt 1
        # for col, val in enumerate(text, start=1):
        #     worksheet.cell(row=2, column=col).value = val

        # Attempt 2 
        # for row_num, data in enumerate(text):
        #     worksheet.write_row(row_num, 0, data)

        # Attempt 3
        # print(list(text[1]))
        for i in range(0,len(text)):
            for j in range(0,len(text[i])):
                worksheet.cell(row=i+1,column=j+1).value = list(text[i])[j]
        
        # close work book
        workbook.save(path)

    def clear_Excel(path, placeholder="Placeholder"):

        workbook  = xlsxwriter.Workbook(path)
        worksheet = workbook.add_worksheet(placeholder)
        workbook.close()

        # workbook = openpyxl.load_workbook(path)
        # print("=" * 100)
        # workbook.create_sheet(placeholder)
        

        # for wk_sheet in workbook.sheetnames: 
        #     ws = workbook[wk_sheet]
        #     workbook.remove(ws)

        # workbook.save(path)

        return

    def save_toExcel_Jyxx_list(text_list,path,wk_sheet="main"):
        text = []
        for item in text_list:
            item = item.values()
            text.append(item)
        # print(text)
        text = [['发布时间','标题','链接']] +  text
        save_toExcel(text,path=path,wk_sheet=wk_sheet)

    def save_toExcel_Jyxx_dict(text_dict,path):
        for item in text_dict.items():
            save_toExcel_Jyxx_list(item[1],wk_sheet=item[0],path=path)

